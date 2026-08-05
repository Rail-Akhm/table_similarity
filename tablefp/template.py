"""Load xlsx template and hash values through the database."""

import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell import Cell
import numpy as np
import psycopg2

from tablefp.hashing import build_ngram_hashes
from tablefp.norm import get_h64_expr, get_norm_expr

logger = logging.getLogger(__name__)


@dataclass
class TemplateColumn:
    """A column from the template."""

    name: str
    values: List[Any]  # Raw cell values
    dtype_group: str
    distinct_hashes: np.ndarray
    row_hashes: List[Optional[int]] = field(default_factory=list)
    row_norm_v: List[Optional[str]] = field(default_factory=list)
    ngram_hashes: np.ndarray = field(default_factory=lambda: np.array([], dtype=np.int64))
    min_val: Optional[float] = None
    max_val: Optional[float] = None


@dataclass
class Template:
    """Loaded xlsx template."""

    path: str
    columns: List[TemplateColumn] = field(default_factory=list)
    header_row: Optional[int] = None


def canonicalize_cell(value: Any) -> Optional[str]:
    """Canonicalize a cell value to text for DB normalization.

    This is minimal preprocessing - the DB will do the real normalization.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        # Up to 6 decimals, strip trailing zeros
        s = f"{value:.6f}".rstrip("0").rstrip(".")
        return s
    if isinstance(value, int):
        return str(value)
    if hasattr(value, "strftime"):
        # datetime or date
        if hasattr(value, "hour"):
            return value.strftime("%Y-%m-%dT%H:%M:%S")
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value.strip() if value.strip() else None
    return str(value)


def infer_dtype_group(values: List[Any]) -> str:
    """Infer dtype group from raw cell values."""
    num_count = 0
    date_count = 0
    total = 0

    for v in values:
        if v is None or v == "":
            continue
        total += 1

        # Try to parse as number
        try:
            float(v)
            num_count += 1
            continue
        except (ValueError, TypeError):
            pass

        # Try to parse as date
        if isinstance(v, str):
            if "T" in v and len(v) >= 19:  # YYYY-MM-DDTHH:MM:SS
                date_count += 1
                continue
            if len(v) == 10 and v[4] == "-" and v[7] == "-":  # YYYY-MM-DD
                date_count += 1
                continue

    if total == 0:
        return "text"

    if num_count / total >= 0.9:
        return "num"
    if date_count / total >= 0.9:
        return "date"

    return "text"


def load_template(
    path: str,
    conn: psycopg2.extensions.connection,
    sheet_name: Optional[str] = None,
    header: Optional[bool] = None,
    min_template_distinct: int = 5,
    fuzzy_enabled: bool = False,
    ngram_size: int = 3,
) -> Template:
    """Load an xlsx template file.

    Args:
        path: Path to xlsx file
        conn: Database connection for hashing
        sheet_name: Sheet name (default: first sheet)
        header: If True, first row is header. If False, no header.
                If None, auto-detect.

    Returns:
        Template with columns and their hashed values
    """
    logger.debug(f"Loading xlsx file: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if sheet_name:
        logger.debug(f"Using sheet: {sheet_name}")
        ws = wb[sheet_name]
    else:
        logger.debug("Using active sheet")
        ws = wb.active

    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    logger.debug(f"Read {len(rows)} rows")

    if not rows:
        raise ValueError("Empty template file")

    # Auto-detect header if needed
    if header is None:
        # Check if first row is all non-numeric strings
        first_row = rows[0]
        header = all(
            isinstance(v, str) and not v.strip().isdigit()
            for v in first_row
            if v is not None
        )

    start_row = 1 if header else 0
    header_names = rows[0] if header else [f"col_{i}" for i in range(len(rows[0]))]

    # Transpose to get columns
    num_cols = len(rows[0])
    columns_data = [[] for _ in range(num_cols)]

    for row in rows[start_row:]:
        for i, v in enumerate(row):
            if i < num_cols:
                # Keep None placeholders so all columns stay row-aligned
                columns_data[i].append(canonicalize_cell(v))

    # Filter columns with >= min_template_distinct non-null values
    template_columns = []

    for i, (name, values) in enumerate(zip(header_names, columns_data)):
        n_present = sum(1 for v in values if v is not None)
        if n_present < min_template_distinct:
            logger.debug(f"  skip '{name}': only {n_present} values (need {min_template_distinct})")
            continue

        logger.debug(f"Processing column '{name}' ({len(values)} values)")

        # Infer dtype
        dtype_group = infer_dtype_group(values)
        logger.debug(f"  Inferred dtype: {dtype_group}")

        # Hash through DB - use 'v' as the column name in the expressions
        norm_expr = get_norm_expr(dtype_group, "v")
        h64_expr = get_h64_expr(dtype_group, "v")

        query = f"""
            SELECT col_idx, {norm_expr} AS norm_v, {h64_expr} AS h
            FROM unnest(%(cols)s::int[], %(vals)s::text[]) AS t(col_idx, v)
        """

        cols_with_idx = []
        vals_with_idx = []
        for row_idx, val in enumerate(values):
            cols_with_idx.append(row_idx)
            vals_with_idx.append(val)

        with conn.cursor() as cur:
            cur.execute(query, {"cols": cols_with_idx, "vals": vals_with_idx})
            db_results = {row[0]: (row[1], row[2]) for row in cur.fetchall()}

        # Per-row hashes and normalized values
        row_hashes = []
        row_norm_v = []
        all_col_hashes = set()

        for row_idx, val in enumerate(values):
            result = db_results.get(row_idx)
            if result is not None and result[1] is not None:
                nv, h = result
                row_hashes.append(h)
                row_norm_v.append(nv)
                all_col_hashes.add(h)
            else:
                row_hashes.append(None)
                row_norm_v.append(None)

        # Get stats for num columns
        min_val = max_val = None
        if dtype_group == "num":
            try:
                numeric_vals = [float(v) for v in values if v is not None]
                if numeric_vals:
                    min_val = min(numeric_vals)
                    max_val = max(numeric_vals)
                    logger.debug(f"  Numeric range: {min_val} - {max_val}")
            except (ValueError, TypeError):
                pass

        # Compute n-gram hashes for text columns when fuzzy is enabled
        ngram_hashes = np.array([], dtype=np.int64)
        if fuzzy_enabled and dtype_group == "text":
            norm_values = [v for v in row_norm_v if v is not None]
            if norm_values:
                ngram_hashes = build_ngram_hashes(norm_values, n=ngram_size)
                logger.debug(f"  {len(ngram_hashes)} n-gram hashes")

        template_columns.append(
            TemplateColumn(
                name=name,
                values=values,
                dtype_group=dtype_group,
                distinct_hashes=np.sort(np.array(list(all_col_hashes), dtype=np.int64)),
                row_hashes=row_hashes,
                row_norm_v=row_norm_v,
                ngram_hashes=ngram_hashes,
                min_val=min_val,
                max_val=max_val,
            )
        )
        logger.debug(f"  {len(all_col_hashes)} distinct hashes")

    logger.debug(f"Loaded {len(template_columns)} columns from template")
    return Template(path=path, columns=template_columns)


def load_raw_columns(
    path: str,
    sheet_name: Optional[str] = None,
    header: Optional[bool] = None,
) -> List[TemplateColumn]:
    """Load every template column as raw values only (no DB hashing).

    Used for display of non-matched columns in `--columns all` comparison. Each
    returned TemplateColumn has `name` and `values` populated; hashing/norm
    fields are left empty. Row order matches load_template (same iteration).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    if header is None:
        first_row = rows[0]
        header = all(
            isinstance(v, str) and not v.strip().isdigit()
            for v in first_row if v is not None
        )

    start_row = 1 if header else 0
    num_cols = len(rows[0])
    header_names = rows[0] if header else [f"col_{i}" for i in range(num_cols)]

    columns_data = [[] for _ in range(num_cols)]
    for row in rows[start_row:]:
        for i, v in enumerate(row):
            if i < num_cols:
                columns_data[i].append(canonicalize_cell(v))

    out = []
    for name, values in zip(header_names, columns_data):
        out.append(TemplateColumn(
            name=name, values=values, dtype_group="text",
            distinct_hashes=np.array([], dtype=np.int64),
        ))
    return out


def hash_template_values(
    conn: psycopg2.extensions.connection,
    values: List[str],
    dtype_group: str,
) -> np.ndarray:
    """Hash a list of template values through the database.

    Returns sorted array of unique hashes.
    """
    if not values:
        return np.array([], dtype=np.int64)

    h64_expr = get_h64_expr(dtype_group, "v")

    query = f"SELECT DISTINCT {h64_expr} FROM unnest(%s::text[]) AS t(v)"

    with conn.cursor() as cur:
        cur.execute(query, (values,))
        hashes = [row[0] for row in cur.fetchall() if row[0] is not None]

    return np.sort(np.array(hashes, dtype=np.int64))